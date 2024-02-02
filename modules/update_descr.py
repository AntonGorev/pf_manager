import os
from yahoofinancials import YahooFinancials
#import yahoo_fin as yf
import yahoo_fin.stock_info as si
import yfinance as yf
import pandas as pd
import numpy as np
from tqdm import tqdm
import math
import time 
from openpyxl import load_workbook
from datetime import datetime
from wakepy import set_keepawake, unset_keepawake

multiplier = {"K": 1000,
                  "M": 1000000,
                  "B": 1000000000,
                  "T": 1000000000000}

path = 'myrics_main_highcap_filtered3.xlsx'
#path = 'myrics_main_light3.xlsx'
#path = 'myrics_main_light3_processed.xlsx'

df = pd.read_excel(path, index_col=0, sheet_name='Main')

i = 0

#tickers = list(df[df['Interesting'].notnull()].index)
tickers = list(df[df['Update Status']=="#to_update"].index) 


#tickers = list(df[df['Update Status'].isnull()].index) 
#tickers = list(df.index)
#tickers = list(df[df['Interesting']=="b-corp"].index) 
print(len(tickers[:]))

#tickers = list(df[df['Industry']!="#bad"].index)
#tickers = list(df[df['Market Cap'].isnull()].index)
#tickers.extend(list(df[df['Comment'].notnull()][df[df['Comment'].notnull()]['Update Status'].isnull()].index))

set_keepawake(keep_screen_awake=False)

for tick in tqdm(tickers[:]):
    print(tick)

    insinfo = yf.Ticker(tick).info
    try:
        if pd.isnull(df.loc[tick, 'Industry']):
            df.loc[tick, 'Industry'] = insinfo['industry']
            df.loc[tick, 'Name2'] = insinfo['longName']
            df.loc[tick, 'Update Status'] = "Done"
    except:
        df.loc[tick, 'Industry'] = "#bad"

    # some more additional data
    try:
        df.loc[tick, 'exchange'] = insinfo['exchange']
        df.loc[tick, 'sector'] = insinfo['sector']
        df.loc[tick, 'earningsGrowth'] = insinfo['earningsGrowth']
        df.loc[tick, 'longBusinessSummary'] = insinfo['longBusinessSummary']
        df.loc[tick, 'Currency'] = insinfo['currency']
        df.loc[tick, 'Website'] = insinfo['website']
        df.loc[tick, 'isEsgPopulated'] = insinfo['isEsgPopulated']
        df.loc[tick, 'beta3Year'] = insinfo['beta3Year']
        df.loc[tick, '52WeekChange'] = insinfo['52WeekChange']
    except:
        pass 

    i += 1
    if i%200==0 or i>=len(tickers):
        df.to_csv('main_draft_light.csv', header=True, index_label=True)

    if i%300==0 or i>=len(tickers):
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book

        # delet existing "Main" sheet
        std = book['Main']
        book.remove(std)

        # write df to a new Main sheet
        df.to_excel(writer, sheet_name='Main')
        #df_pf.to_excel(writer, sheet_name='pf')
        writer.save()
        writer.close()

    time.sleep(0.5)

df["Cheap Asset"] = df["Market Cap"] <= df["Cheap Assets (1.5(TA-TL))"] 

#df_pf = df[df['Comment'].notnull()]
#df_pf = df[df['Portfolio']==1]

# write to workbook
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book

# delet existing "Main" sheet
std = book['Main']
book.remove(std)

# write df to a new Main sheet
df.to_excel(writer, sheet_name='Main')
#df_pf.to_excel(writer, sheet_name='pf')
writer.save()
writer.close()

unset_keepawake()

#################
#for key, value in insinfo.items() :
#    print(key, value)