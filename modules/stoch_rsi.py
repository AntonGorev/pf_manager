import os
import sys
from yahoofinancials import YahooFinancials
#import yahoo_fin as yf
import yahoo_fin.stock_info as si
import yfinance as yf
import pandas as pd
import numpy as np
from tqdm import tqdm
import math
import time 
from openpyxl import load_workbook
from datetime import datetime
from wakepy import set_keepawake, unset_keepawake
import py_project.modules.plotting as myplt
from datetime import datetime, timedelta

###############################################################
def computeRSI (data, time_window):
    diff = data.diff(1).dropna()        # diff in one field(one day)

    #this preservers dimensions off diff values
    up_chg = 0 * diff
    down_chg = 0 * diff
    
    # up change is equal to the positive difference, otherwise equal to zero
    up_chg[diff > 0] = diff[ diff>0 ]
    
    # down change is equal to negative deifference, otherwise equal to zero
    down_chg[diff < 0] = diff[ diff < 0 ]
    
    # check pandas documentation for ewm
    # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.ewm.html
    # values are related to exponential decay
    # we set com=time_window-1 so we get decay alpha=1/time_window
    up_chg_avg   = up_chg.ewm(com=time_window-1 , min_periods=time_window).mean()
    down_chg_avg = down_chg.ewm(com=time_window-1 , min_periods=time_window).mean()
    
    rs = abs(up_chg_avg/down_chg_avg)
    rsi = 100 - 100/(1+rs)
    return rsi

###############################################################
def stochastic(data, k_window, d_window, window):
    
    # input to function is one column from df
    # containing closing price or whatever value we want to extract K and D from
    
    min_val  = data.rolling(window=window, center=False).min()
    max_val = data.rolling(window=window, center=False).max()
    
    stoch = ( (data - min_val) / (max_val - min_val) ) * 100
    
    K = stoch.rolling(window=k_window, center=False).mean() 
    #K = stoch
    
    D = K.rolling(window=d_window, center=False).mean() 


    return K, D

def signal(k, d):
    try:
        if k > 80 and k<=d:
            s = "sell"
        elif k > 80 and k>d:
            s="sell_wait"
        elif k < 20 and k>=d:
            s="buy"
        elif k < 20 and k<d:
            s="buy_wait"
        else:
            s="wait"
    except:
        s=None

    return s

################# start ###################
path = 'py_project/input/myrics_main_highcap_filtered3.xlsx'

df_rics = pd.read_excel(path, index_col=0, sheet_name='Recs')

df_rics['RSI_W'] = None
df_rics['K_W'] = None
df_rics['D_W'] = None
df_rics['RSI_D'] = None
df_rics['K_D'] = None
df_rics['D_D'] = None

tickers = list(df_rics.index)
print(len(tickers[:]))

end_d = datetime.now().strftime("%Y-%m-%d")
start_d = datetime.now() - timedelta(days=365)
start_d = start_d.strftime("%Y-%m-%d")

agg_dict = {'Open': 'first',
          'High': 'max',
          'Low': 'min',
          'Close': 'last',
          'Adj Close': 'last',
          'Volume': 'mean'}

set_keepawake(keep_screen_awake=False)

for tick in tqdm(tickers[:]):
    print(tick)
    try:
        df = yf.download(tick, start=start_d, end=end_d, interval = "1d")
        df['RSI'] = computeRSI(df['Adj Close'], 14)
        df['K'], df['D'] = stochastic(df['RSI'], 3, 3, 14)
        
        for period in ['W', 'D']:
            try:
                if period == 'W':
                    df_res = df.resample(period).agg(agg_dict)
                else:
                    df_res=df

                df_res['RSI'] = computeRSI(df_res['Adj Close'], 14)
                df_res['K'], df_res['D'] = stochastic(df_res['RSI'], 3, 3, 14)

                df_rics.loc[tick, 'RSI_'+period] = df_res['RSI'].iloc[-1]
                df_rics.loc[tick, 'K_'+period] = df_res['K'].iloc[-1]
                df_rics.loc[tick, 'D_'+period] = df_res['D'].iloc[-1]
                df_rics.loc[tick, 'Stoch_RSI_'+period] = signal(df_res['K'].iloc[-1], df_res['D'].iloc[-1])
            except:
                df_rics.loc[tick, 'RSI_'+period] = None
                df_rics.loc[tick, 'K_'+period] = None
                df_rics.loc[tick, 'D_'+period] = None
                df_rics.loc[tick, 'Stoch_RSI_'+period] = None

            #df_rics.loc[df_rics['Ticker'] == tick, 'D'] = df_res['D'].iloc[-1]
    except:
        pass

    time.sleep(1)

# write to workbook
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book
# book.sheetnames

# delet existing "Main" sheet
try:
    std = book['Recs_python']
    book.remove(std)
except:
    pass

# write df to a new Main sheet
df_rics.to_excel(writer, sheet_name='Recs_python')
writer.save()
writer.close()

unset_keepawake()
